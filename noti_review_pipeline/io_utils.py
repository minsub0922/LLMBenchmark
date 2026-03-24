import csv
import json
import os
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def ensure_dir(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path


def load_jsonl(path: str) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception as e:
                raise ValueError(f"Invalid JSONL at line {line_no}: {e}") from e
            if not isinstance(obj, dict):
                raise ValueError(f"Line {line_no} is not a JSON object")
            records.append(obj)
    return records


def dump_jsonl(path: str, records: List[Dict[str, Any]]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def dump_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    fieldnames = sorted({k for row in rows for k in row.keys()}) if rows else []
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _autosize(ws, max_width: int = 80) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = min(max(widths.get(cell.column_letter, 0), len(str(cell.value)) + 2), max_width)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def dump_xlsx(path: str, sheets: Dict[str, List[Dict[str, Any]]]) -> None:
    wb = Workbook()
    first = True
    fills = ["D9EAF7", "FCE4D6", "E2F0D9", "FFF2CC", "EADCF8"]

    for idx, (sheet_name, rows) in enumerate(sheets.items()):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_name[:31]
        headers = sorted({k for row in rows for k in row.keys()}) if rows else []
        if headers:
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor=fills[idx % len(fills)])
                c.alignment = Alignment(vertical="top", wrap_text=True)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
        else:
            ws.append(["message"])
            ws.append(["no data"])
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _autosize(ws)

    wb.save(path)
