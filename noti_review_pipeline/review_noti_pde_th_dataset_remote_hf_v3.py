#!/usr/bin/env python3
"""
Human-feedback-aware remote review runner v3.

What is added on top of v2:
- detailed per-row change history
- explicit change type columns for filtering in Excel
- title-before / title-after tracking
- body-before / body-after tracking
- reason columns based on model issues
- workbook sheets for full history, changed-only rows, summary, final records
- Excel auto-filter enabled on every sheet

Interface is intentionally kept compatible with the previous remote HF runner style.
"""

import argparse
import json
import os
import re
import sys
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
if CURRENT_DIR not in sys.path:
    sys.path.insert(0, CURRENT_DIR)

from config import DEFAULT_BATCH_SIZE, DEFAULT_ENDPOINT, DEFAULT_MAX_PASSES, DEFAULT_MODEL, DEFAULT_MODE
from io_utils import dump_csv, dump_jsonl, ensure_dir, load_jsonl
from remote_client import RemoteInferenceClient


HUMAN_FEEDBACK_PROFILES: Dict[str, str] = {
    "none": "",
    "title_grammar_v1": (
        "Apply these human-feedback-based preferences only when clearly supported by the sample text:\n"
        "1. If the title is a weak synthetic label, duplicated brand phrase, or generic confirmation phrase, rewrite the title to a more useful reservation-style title.\n"
        "2. Prefer the Thai title pattern 'การจอง {hotel_name} {room_type}' when both hotel/property name and room type are explicitly present in the sample.\n"
        "3. Remove duplicated booking-brand prefixes or generic phrases like 'การยืนยันการจอง' when a clearer reservation title can be formed.\n"
        "4. Improve wrong grammar, awkward Thai phrasing, spacing, and punctuation in the raw notification text.\n"
        "5. Do not invent hotel name, room type, date, or other facts that are not explicitly present.\n"
        "6. Keep semantic meaning, reservation status, ids, and all factual content unchanged.\n"
    ),
    "title_content_coverage_v1": (
        "Apply these title-content-coverage preferences only when clearly supported by the sample text:\n"
        "1. If the title is too generic, such as only 'การจอง', 'การจองของคุณได้รับการยืนยัน', or similar, rewrite it so the main reserved entity is visible in the title.\n"
        "2. The rewritten title should cover the core content of the reservation, especially the venue / property / service name explicitly shown in the sample.\n"
        "3. Prefer the compact Thai pattern 'การจอง {main_entity}' when the sample mainly provides one clear reserved entity, for example an event space, venue, hotel, or service name.\n"
        "4. If both a clear property/service name and a clear room/item name are explicitly present, you may prefer 'การจอง {property_name} {room_or_item_name}' only when it remains natural and not overly long.\n"
        "5. Do not leave the title as a bare generic reservation label if the body clearly identifies what was reserved.\n"
        "6. Do not invent missing entities. Keep all factual content and meaning unchanged.\n"
    ),
    "title_grammar_coverage_v1": (
        "Combine grammar cleanup and title-content coverage rules when clearly supported by the sample text:\n"
        "1. Improve wrong grammar, awkward Thai phrasing, spacing, punctuation, and malformed line breaks in the raw notification text.\n"
        "2. If the title is generic or summary-like and does not cover the actual reservation content, rewrite it to expose the main reserved entity.\n"
        "3. Prefer 'การจอง {main_entity}' when the sample provides one clear venue/property/service name.\n"
        "4. Prefer 'การจอง {hotel_name} {room_type}' when both hotel/property and room/item are explicitly present and the result is natural.\n"
        "5. Remove duplicated brand prefixes or generic confirmation wording when a clearer reservation title can be formed.\n"
        "6. Do not invent facts. Keep reservation status, ids, dates, times, entities, and semantics unchanged.\n"
    ),
}


SYSTEM_RULES = """
You are a careful dataset quality reviewer for reservation notification extraction data.

Your goal is NOT to change the meaning, facts, labels, entity values, dates, times, addresses, reservation ids, or cancellation/confirmation status.
You may ONLY improve data quality in limited ways:
1. fix obvious grammar or spacing issues in the raw notification text
2. fix broken punctuation or malformed line breaks
3. improve title text when it is a low-quality synthetic summary label rather than a natural title, but preserve meaning
4. normalize tiny expression issues without changing semantics
5. never invent missing facts
6. never rewrite the sample into a different style if the original is already acceptable

Return strict JSON:
{
  "decision": "keep" | "improve",
  "issues": ["..."],
  "improved_text": "...",
  "quality_scores": {
    "grammar": 1-5,
    "naturalness": 1-5,
    "title_quality": 1-5,
    "format_consistency": 1-5
  }
}
""".strip()


TITLE_RE = re.compile(r"(^|\n)Title:\s*(.*)")
SUBTEXT_RE = re.compile(r"(^|\n)Sub[_ ]?text:\s*(.*)", re.IGNORECASE)
MESSAGE_RE = re.compile(r"(^|\n)Message:\s*(.*)", re.IGNORECASE | re.DOTALL)


def build_review_prompt(text: str, checks: List[str], feedback_profile: str = "none") -> str:
    feedback_instruction = HUMAN_FEEDBACK_PROFILES.get(feedback_profile, "")
    prompt = SYSTEM_RULES + "\n\n" + f"Active checks: {', '.join(checks)}\n"
    if feedback_instruction:
        prompt += "\nHuman feedback preferences:\n" + feedback_instruction + "\n"
    prompt += "\nReview this sample and improve it only if needed.\n\n[SAMPLE]\n" + text
    return prompt



def parse_json_response(raw: str) -> Dict[str, Any]:
    raw = raw.strip()
    try:
        return json.loads(raw)
    except Exception:
        pass
    m = re.search(r"\{.*\}", raw, flags=re.DOTALL)
    if not m:
        raise ValueError("No JSON object found in response")
    return json.loads(m.group(0))



def extract_title(text: str) -> str:
    m = TITLE_RE.search(text)
    return m.group(2).strip() if m else ""



def split_text_parts(text: str) -> Dict[str, str]:
    title = extract_title(text)
    sub_text = ""
    message = ""

    m = SUBTEXT_RE.search(text)
    if m:
        sub_text = m.group(2).strip()

    m = MESSAGE_RE.search(text)
    if m:
        message = m.group(2).strip()

    return {
        "title": title,
        "sub_text": sub_text,
        "message": message,
    }



def normalize_for_compare(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()



def replace_text_field(record: Dict[str, Any], new_text: str) -> Dict[str, Any]:
    out = dict(record)
    out["text"] = new_text
    return out



def detect_change_tags(issues: List[str], title_changed: bool, body_changed: bool) -> List[str]:
    tags: List[str] = []
    lowered = " | ".join(issues).lower() if issues else ""

    if title_changed:
        tags.append("title_changed")
    if body_changed:
        tags.append("body_changed")

    if any(x in lowered for x in ["grammar", "wrong grammar", "awkward", "unnatural"]):
        tags.append("grammar_fix")
    if any(x in lowered for x in ["spacing", "punctuation", "line break", "format", "formatting"]):
        tags.append("format_fix")
    if any(x in lowered for x in ["summary label", "generic title", "title", "coverage"]):
        tags.append("title_quality_fix")

    if not tags:
        tags.append("other")
    return sorted(set(tags))



def detect_primary_change_type(changed: bool, title_changed: bool, body_changed: bool, tags: List[str]) -> str:
    if not changed:
        return "no_change"
    if title_changed and not body_changed:
        return "title_only"
    if body_changed and not title_changed:
        if "format_fix" in tags and "grammar_fix" not in tags:
            return "format_only"
        if "grammar_fix" in tags and "format_fix" not in tags:
            return "grammar_only"
        return "body_only"
    if title_changed and body_changed:
        return "title_and_body"
    return "other_change"



def build_history_row(
    row_id: int,
    pass_idx: int,
    feedback_profile: str,
    original_text: str,
    final_text: str,
    decision: str,
    changed: bool,
    issues: List[str],
    scores: Dict[str, Any],
) -> Dict[str, Any]:
    before_parts = split_text_parts(original_text)
    after_parts = split_text_parts(final_text)

    title_before = before_parts["title"]
    title_after = after_parts["title"]

    title_changed = normalize_for_compare(title_before) != normalize_for_compare(title_after)
    body_before = normalize_for_compare(original_text.replace(title_before, "", 1) if title_before else original_text)
    body_after = normalize_for_compare(final_text.replace(title_after, "", 1) if title_after else final_text)
    body_changed = body_before != body_after

    change_tags = detect_change_tags(issues, title_changed, body_changed)
    change_type = detect_primary_change_type(changed, title_changed, body_changed, change_tags)
    changed_fields = []
    if title_changed:
        changed_fields.append("title")
    if body_changed:
        changed_fields.append("body")

    return {
        "row_id": row_id,
        "pass_idx": pass_idx,
        "feedback_profile": feedback_profile,
        "decision": decision,
        "changed": changed,
        "change_type": change_type,
        "change_tags": ", ".join(change_tags),
        "changed_fields": ", ".join(changed_fields) if changed_fields else "",
        "reason": " | ".join(issues) if issues else "",
        "title_before": title_before,
        "title_after": title_after,
        "title_changed": title_changed,
        "body_changed": body_changed,
        "grammar_score": scores.get("grammar", ""),
        "naturalness_score": scores.get("naturalness", ""),
        "title_quality_score": scores.get("title_quality", ""),
        "format_consistency_score": scores.get("format_consistency", ""),
        "original_text": original_text,
        "final_text": final_text,
    }



def decide_and_apply(
    record: Dict[str, Any],
    review: Dict[str, Any],
    dry_run: bool,
    row_id: int,
    pass_idx: int,
    feedback_profile: str,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    original_text = str(record.get("text", ""))
    decision = str(review.get("decision", "keep")).strip().lower()
    improved_text = str(review.get("improved_text", original_text))
    issues = review.get("issues", [])
    if not isinstance(issues, list):
        issues = [str(issues)]
    scores = review.get("quality_scores", {})
    if not isinstance(scores, dict):
        scores = {}

    changed = decision == "improve" and improved_text.strip() and normalize_for_compare(improved_text) != normalize_for_compare(original_text) and not dry_run
    updated_record = replace_text_field(record, improved_text) if changed else dict(record)
    final_text = str(updated_record.get("text", ""))

    history_row = build_history_row(
        row_id=row_id,
        pass_idx=pass_idx,
        feedback_profile=feedback_profile,
        original_text=original_text,
        final_text=final_text,
        decision=decision,
        changed=changed,
        issues=issues,
        scores=scores,
    )
    return updated_record, history_row



def autosize_worksheet(ws) -> None:
    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            value_str = "" if value is None else str(value)
            width = min(max(len(value_str), 10), 80)
            max_widths[idx] = max(max_widths.get(idx, 10), width)
    for idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width



def write_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    headers = sorted({k for row in rows for k in row.keys()}) if rows else ["message"]
    ws.append(headers)
    if rows:
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    else:
        ws.append(["no data"] + [""] * (len(headers) - 1))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_worksheet(ws)



def save_history_workbook(path: str, history_rows: List[Dict[str, Any]], final_records: List[Dict[str, Any]]) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "history_all"
    write_sheet(ws1, history_rows)

    ws2 = wb.create_sheet("changed_only")
    write_sheet(ws2, [r for r in history_rows if r.get("changed")])

    ws3 = wb.create_sheet("change_summary")
    change_counter = Counter(str(r.get("change_type", "")) for r in history_rows)
    tag_counter = Counter()
    for row in history_rows:
        for tag in str(row.get("change_tags", "")).split(","):
            tag = tag.strip()
            if tag:
                tag_counter[tag] += 1
    summary_rows = [{"metric": "total_rows", "value": len(history_rows)}]
    summary_rows += [{"metric": f"change_type::{k}", "value": v} for k, v in sorted(change_counter.items())]
    summary_rows += [{"metric": f"change_tag::{k}", "value": v} for k, v in sorted(tag_counter.items())]
    write_sheet(ws3, summary_rows)

    ws4 = wb.create_sheet("final_records")
    write_sheet(ws4, final_records)

    wb.save(path)



def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input_file", type=str, required=True)
    parser.add_argument("--output_dir", type=str, required=True)
    parser.add_argument("--endpoint", type=str, default=os.environ.get("ENDPOINT", DEFAULT_ENDPOINT))
    parser.add_argument("--model", type=str, default=os.environ.get("MODEL", DEFAULT_MODEL))
    parser.add_argument("--workers", type=int, default=int(os.environ.get("WORKERS", "4")))
    parser.add_argument("--timeout", type=int, default=int(os.environ.get("TIMEOUT", "180")))
    parser.add_argument("--max_retries", type=int, default=int(os.environ.get("MAX_RETRIES", "3")))
    parser.add_argument("--retry_sleep", type=float, default=float(os.environ.get("RETRY_SLEEP", "2.0")))
    parser.add_argument("--batch_size", type=int, default=int(os.environ.get("BATCH_SIZE", str(DEFAULT_BATCH_SIZE))))
    parser.add_argument("--max_passes", type=int, default=int(os.environ.get("MAX_PASSES", str(DEFAULT_MAX_PASSES))))
    parser.add_argument("--mode", type=str, default=os.environ.get("MODE", DEFAULT_MODE))
    parser.add_argument("--dry_run", action="store_true", default=os.environ.get("DRY_RUN", "0") == "1")
    parser.add_argument("--checks", type=str, default=os.environ.get("CHECKS", "grammar,title_summary,field_consistency"))
    parser.add_argument("--feedback_profile", type=str, default=os.environ.get("FEEDBACK_PROFILE", "none"))
    return parser.parse_args()



def run_slug(checks: List[str], max_passes: int, mode: str, dry_run: bool, feedback_profile: str) -> str:
    checks_slug = "-".join(checks) if checks else "none"
    fb_slug = f"__fb_{feedback_profile}" if feedback_profile and feedback_profile != "none" else ""
    return f"{checks_slug}__p{max_passes}__{mode}{'__dry' if dry_run else ''}{fb_slug}"



def review_one(
    client: RemoteInferenceClient,
    idx: int,
    record: Dict[str, Any],
    pass_idx: int,
    checks: List[str],
    dry_run: bool,
    feedback_profile: str,
) -> Tuple[int, Dict[str, Any], Dict[str, Any]]:
    prompt = build_review_prompt(str(record.get("text", "")), checks, feedback_profile=feedback_profile)
    raw = client.generate(prompt)
    review = parse_json_response(raw)
    updated, row = decide_and_apply(
        record=record,
        review=review,
        dry_run=dry_run,
        row_id=idx,
        pass_idx=pass_idx,
        feedback_profile=feedback_profile,
    )
    return idx, updated, row



def review_pass(
    client: RemoteInferenceClient,
    records: List[Dict[str, Any]],
    pass_idx: int,
    workers: int,
    checks: List[str],
    dry_run: bool,
    feedback_profile: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    next_records: List[Dict[str, Any]] = [None] * len(records)  # type: ignore
    rows: List[Dict[str, Any]] = [None] * len(records)  # type: ignore

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [
            ex.submit(review_one, client, idx, rec, pass_idx, checks, dry_run, feedback_profile)
            for idx, rec in enumerate(records)
        ]
        for fut in as_completed(futures):
            idx, updated, row = fut.result()
            next_records[idx] = updated
            rows[idx] = row

    return next_records, rows



def main():
    args = parse_args()
    ensure_dir(args.output_dir)

    checks = [x.strip() for x in args.checks.split(",") if x.strip()]
    client = RemoteInferenceClient(
        endpoint=args.endpoint,
        model=args.model,
        timeout=args.timeout,
        max_retries=args.max_retries,
        retry_sleep=args.retry_sleep,
    )

    print(f"[INFO] endpoint={args.endpoint}")
    print(f"[INFO] model={args.model}")
    print(f"[INFO] workers={args.workers}")
    print(f"[INFO] batch_size={args.batch_size}")
    print(f"[INFO] checks={checks}")
    print(f"[INFO] max_passes={args.max_passes}")
    print(f"[INFO] mode={args.mode}")
    print(f"[INFO] feedback_profile={args.feedback_profile}")

    records = load_jsonl(args.input_file)
    base_name = os.path.splitext(os.path.basename(args.input_file))[0]
    slug = run_slug(checks, args.max_passes, args.mode, args.dry_run, args.feedback_profile)

    history_rows: List[Dict[str, Any]] = []
    current_records = records

    if args.mode in {"full", "review_only", "apply_only"}:
        for pass_idx in range(1, args.max_passes + 1):
            current_records, rows = review_pass(
                client=client,
                records=current_records,
                pass_idx=pass_idx,
                workers=args.workers,
                checks=checks,
                dry_run=args.dry_run,
                feedback_profile=args.feedback_profile,
            )
            history_rows.extend(rows)
            if args.mode == "review_only":
                break

    improved_jsonl = os.path.join(args.output_dir, f"{base_name}__{slug}.improved.jsonl")
    history_csv = os.path.join(args.output_dir, f"{base_name}__{slug}.history.csv")
    history_xlsx = os.path.join(args.output_dir, f"{base_name}__{slug}.history.xlsx")

    dump_jsonl(improved_jsonl, current_records)
    dump_csv(history_csv, history_rows)
    save_history_workbook(history_xlsx, history_rows, current_records)

    print("[INFO] Done")
    print(f"[INFO] improved_jsonl={improved_jsonl}")
    print(f"[INFO] history_csv={history_csv}")
    print(f"[INFO] history_xlsx={history_xlsx}")


if __name__ == "__main__":
    main()
